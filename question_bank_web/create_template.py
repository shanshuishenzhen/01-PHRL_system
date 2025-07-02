#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
题库模板生成脚本
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_question_bank_template():
    """创建题库模板"""
    
    # 创建示例数据
    template_data = {
        '题库名称': ['保卫管理员（三级）理论']*4,
        'ID': ['B-A-B-C-001-001', 'G-A-B-C-001-002', 'C-A-B-C-001-003', 'T-A-B-C-001-004'],
        '序号': ['1', '2', '3', '4'],
        '认定点代码': ['A-B-C-001', 'A-B-C-001', 'A-B-C-001', 'A-B-C-001'],
        '题型代码': ['B（单选题）', 'G（多选题）', 'C（判断题）', 'T（填空题）'],
        '题号': ['001', '002', '003', '004'],
        '试题（题干）': [
            '示例：下列哪项属于安全管理的基本原则？',
            '示例：关于健康管理，下列哪些说法是正确的？',
            '示例：健康管理师的主要职责包括健康教育。',
            '示例：请填写健康管理的核心内容。'
        ],
        '试题（选项A）': ['A. 预防为主', 'A. 仅体检', '', ''],
        '试题（选项B）': ['B. 以罚代管', 'B. 个体化干预', '', ''],
        '试题（选项C）': ['C. 事后补救', 'C. 过程管理', '', ''],
        '试题（选项D）': ['D. 只重结果', 'D. 结果评估', '', ''],
        '试题（选项E）': ['', 'E. 全程管理', '', ''],
        '【图】及位置': ['', '', '', ''],
        '正确答案': ['A', 'A,B,C', '正确', '健康评估、健康干预、健康教育'],
        '难度代码': ['3（中等）', '4（困难）', '2（简单）', '3（中等）'],
        '一致性代码': ['4（高）', '5（很高）', '3（中等）', '4（高）'],
        '解析': [
            '安全管理应以预防为主。',
            '健康管理包括个体化干预、过程管理等。',
            '健康教育是健康管理师的重要职责。',
            '健康管理的核心内容包括健康评估、干预和教育。'
        ]
    }
    
    # 创建DataFrame，确保题库名称为第一列
    columns = ['题库名称', 'ID', '序号', '认定点代码', '题型代码', '题号', '试题（题干）', '试题（选项A）', '试题（选项B）', '试题（选项C）', '试题（选项D）', '试题（选项E）', '【图】及位置', '正确答案', '难度代码', '一致性代码', '解析']
    df = pd.DataFrame(template_data, columns=columns)
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "题库模板"
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 设置列宽，A列为题库名称
    column_widths = {
        'A': 20,  # 题库名称
        'B': 20,  # ID
        'C': 10,  # 序号
        'D': 15,  # 认定点代码
        'E': 12,  # 题型代码
        'F': 10,  # 题号
        'G': 50,  # 试题（题干）
        'H': 20,  # 试题（选项A）
        'I': 20,  # 试题（选项B）
        'J': 20,  # 试题（选项C）
        'K': 20,  # 试题（选项D）
        'L': 20,  # 试题（选项E）
        'M': 15,  # 【图】及位置
        'N': 15,  # 正确答案
        'O': 12,  # 难度代码
        'P': 12,  # 一致性代码
        'Q': 30   # 解析
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 添加说明工作表
    ws_info = wb.create_sheet("使用说明")
    
    # 添加使用说明
    instructions = [
        ["题库模板使用说明"],
        [""],
        ["1. 文件格式要求"],
        ["   - 文件格式：.xlsx (Excel 2007及以上版本)"],
        ["   - 文件大小：不超过10MB"],
        ["   - 编码格式：UTF-8"],
        [""],
        ["2. 必需列说明"],
        ["   - 题库名称：每道题所属题库，必须与组题规则表中的'题库名称'一致。"],
        ["   - ID：题目唯一标识，格式为 'B-A-B-C-001-002'，共6段，以-分隔。第1段为题型代码（如B），第2-5段为认定点代码（如A-B-C-001），第6段为顺序号（如002）。"],
        ["   - 认定点代码：格式为 'A-B-C-001'，对应ID的第2-5段。"],
        ["   - 题型代码：题目类型，详见下方代码表"],
        ["   - 试题（题干）：题目内容，不能为空"],
        ["   - 正确答案：正确答案内容，不能为空"],
        ["   - 难度代码：难度等级，1-5"],
        [""],
        ["3. 题型代码表"],
        ["   代码", "题型", "说明"],
        ["   B（单选题）", "单选题", "只有一个正确答案"],
        ["   G（多选题）", "多选题", "有多个正确答案，用逗号分隔"],
        ["   C（判断题）", "判断题", "正确/错误"],
        ["   T（填空题）", "填空题", "填写答案内容"],
        ["   D（简答题）", "简答题", "简短回答"],
        ["   U（计算题）", "计算题", "需要计算过程"],
        ["   W（论述题）", "论述题", "详细论述"],
        ["   E（案例分析题）", "案例分析", "案例分析题"],
        ["   F（综合题）", "综合题", "综合多种题型"],
        [""],
        ["4. 难度代码表"],
        ["   代码", "难度", "说明"],
        ["   1（很简单）", "很简单", "基础题目"],
        ["   2（简单）", "简单", "较容易"],
        ["   3（中等）", "中等", "一般难度"],
        ["   4（困难）", "困难", "较难"],
        ["   5（很难）", "很难", "高难度"],
        [""],
        ["5. 一致性代码表"],
        ["   代码", "一致性", "说明"],
        ["   1（很低）", "很低", "一致性很差"],
        ["   2（低）", "低", "一致性较差"],
        ["   3（中等）", "中等", "一般一致性"],
        ["   4（高）", "高", "较好一致性"],
        ["   5（很高）", "很高", "很好一致性"],
        [""],
        ["6. 注意事项"],
        ["   - 请严格按照模板格式填写"],
        ["   - 题型代码必须使用标准代码"],
        ["   - 难度代码和一致性代码范围：1-5"],
        ["   - 选项内容根据题型需要填写"],
        ["   - 图片信息可填写图片路径或描述"],
        [""],
        ["7. 示例数据"],
        ["   模板中已包含4道示例题目，请参考填写格式。"]
    ]
    
    for row in instructions:
        ws_info.append(row)
    
    # 设置说明页面样式
    for row in ws_info.iter_rows():
        for cell in row:
            if cell.value and "代码" in str(cell.value) and "表" in str(cell.value):
                cell.font = Font(bold=True, color="366092")
            elif cell.value and cell.value.startswith(("B", "G", "C", "T", "D", "U", "W", "E", "F", "1", "2", "3", "4", "5")):
                cell.font = Font(bold=True)
    
    # 设置列宽
    ws_info.column_dimensions['A'].width = 15
    ws_info.column_dimensions['B'].width = 20
    ws_info.column_dimensions['C'].width = 30
    
    # 保存文件
    template_path = os.path.join(os.getcwd(), 'templates', '题库模板.xlsx')
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    wb.save(template_path)
    
    print(f"✅ 题库模板已生成: {template_path}")
    return template_path

def main():
    """主函数"""
    print("📋 生成题库模板...")
    try:
        template_path = create_question_bank_template()
        print(f"🎉 模板生成成功！")
        print(f"📁 文件位置: {template_path}")
        print(f"📊 包含示例数据: 4道题目")
        print(f"📖 包含使用说明: 详细的使用指南")
    except Exception as e:
        print(f"❌ 模板生成失败: {e}")

if __name__ == '__main__':
    main()